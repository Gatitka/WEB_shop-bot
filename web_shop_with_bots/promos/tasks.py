from celery import shared_task
from django.utils import timezone
from django.conf import settings

from promos.models import PromoBroadcast
from tm_bot.models import MessengerAccount, MessengerAccountBot
from tm_bot.services import (send_message_telegram,
                             send_user_message_via_bot,
                             send_user_photo_via_bot,
                             build_keyboard_for_broadcast)
from tm_bot.text_assemble_and_edition import clean_html_for_telegram
import time
from urllib.parse import urljoin
from django.db import close_old_connections
from collections import Counter
import json
from django.db import transaction
from django.db.utils import InterfaceError, OperationalError
from django.db.models import Q
from django.db import transaction
from io import BytesIO
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import logging
logger = logging.getLogger("promos")

SAVE_EVERY = 200
# в процессе рассылки каждые 200 контактов сохраняется промежуточный результат


def crop_text_if_photo(cleaned_message, photo_url, broadcast):
    # если есть картинка — caption ограничен 1024 символами
    caption = None
    text_only = None

    if photo_url:
        caption = cleaned_message.strip() or None
        if caption and len(caption) > 1024:
            # обрежем, чтобы не свалиться с ошибкой
            caption = caption[:1018] + "…"
            logger.warning(
                "Broadcast %s: caption truncated to 1024 chars.",
                broadcast.id,
            )
    else:
        text_only = cleaned_message.strip() or None
        if text_only and len(text_only) > 4096:
            text_only = text_only[:4088] + "…"
            logger.warning(
                "Broadcast %s: text truncated to 4096 chars.",
                broadcast.id,
            )
    return caption, text_only


def get_broadcast_photo_url(broadcast: PromoBroadcast) -> str | None:
    """
    Возвращает абсолютный URL до картинки рассылки,
    пригодный для Telegram, либо None, если картинки нет.
    """
    if not broadcast.image:
        return None

    url = broadcast.image.url  # может быть уже абсолютным

    # В production/test_server MEDIA_URL у тебя уже с доменом,
    # а в development — относительный путь /media/...
    if url.startswith("http://") or url.startswith("https://"):
        return url

    base = None
    if getattr(settings, "SERVER", None):
        base = f"{settings.PROTOCOL}://{settings.SERVER}"
    elif getattr(settings, "TEST_SERVER", None):
        base = f"{settings.PROTOCOL}://{settings.TEST_SERVER}"
    else:
        # fallback для локалки — для Telegrаm это работать не будет,
        # но для тестов с ngrok/туннелем ок
        base = "http://localhost:8000"

    return urljoin(base, url)


def build_broadcast_report_file(
    broadcast,
    bot,
    city,
    counter,
    detail_rows,
    *,
    processed_count=None,
    delivered_count=None,
    final_status=None,
):
    wb = Workbook()

    processed_count = processed_count if processed_count is not None else broadcast.processed_count
    delivered_count = delivered_count if delivered_count is not None else broadcast.delivered_count
    final_status = final_status if final_status is not None else broadcast.status

    # ---------------- Sheet 1: summary ----------------
    ws1 = wb.active
    ws1.title = "summary"

    summary_rows = [
        ["broadcast_id", broadcast.id],
        ["title", broadcast.title],
        ["bot_id", bot.id if bot else None],
        ["bot_name", str(bot) if bot else None],
        ["city", city],
        ["status", final_status],
        ["sent_at", str(broadcast.sent_at) if broadcast.sent_at else ""],
        ["total_recipients", broadcast.total_recipients],
        ["processed_count", processed_count],
        ["delivered_count", delivered_count],
    ]

    for row in summary_rows:
        ws1.append(row)

    ws1.append([])
    ws1.append(["metric", "value"])

    for key, value in counter.items():
        ws1.append([key, value])

    # стили для summary
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    if ws1.max_row >= len(summary_rows) + 2:
        for cell in ws1[len(summary_rows) + 2]:
            cell.font = Font(bold=True)

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 40

    # ---------------- Sheet 2: details ----------------
    ws2 = wb.create_sheet("details")
    headers = [
        "messenger_account_id",
        "username",
        "msngr_id",
        "tm_chat_id",
        "city",
        "bot_id",
        "bot_name",
        "status",
        "error_code",
        "description",
        "tg_can_write_before",
        "tg_can_write_after",
        "last_error_code_after",
        "response_json",
    ]
    ws2.append(headers)

    for row in detail_rows:
        ws2.append([row.get(h) for h in headers])

    # стили для details
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ширины колонок
    widths = {
        "A": 18,  # messenger_account_id
        "B": 20,  # username
        "C": 16,  # msngr_id
        "D": 16,  # tm_chat_id
        "E": 14,  # city
        "F": 10,  # bot_id
        "G": 18,  # bot_name
        "H": 20,  # status
        "I": 12,  # error_code
        "J": 40,  # description
        "K": 16,  # tg_can_write_before
        "L": 16,  # tg_can_write_after
        "M": 22,  # last_error_code_after
        "N": 80,  # response_json
    }
    for col, width in widths.items():
        ws2.column_dimensions[col].width = width

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"broadcast_{broadcast.id}_report.xlsx"
    return filename, ContentFile(buffer.read())


@shared_task(
    queue="broadcast")
def send_broadcast_test_task(broadcast_id: int):
    """ Отправка через таску."""
    broadcast, bot, bot_token, city = get_broadcast_bot_token_city(
                                                        broadcast_id)

    # выбираем тестовый админский чат по городу
    #test_chat_id = settings.ADMIN_CHATS.get(city) if city else None

    # выбираем личный чат бота с админом по боту
    test_chat_id = bot.admin_id
    if not test_chat_id or not bot_token:
        return f"No test chat or bot token for city={city}"

    # картинка (может быть None)
    photo_url = get_broadcast_photo_url(broadcast)

    # текст сообщения
    # очищаем HTML форматирование в сообщении
    cleaned_message = clean_html_for_telegram(broadcast.body)
    # обрезаем кол-во символов под стандарт и если есть фото
    caption, text_only = crop_text_if_photo(cleaned_message,
                                            photo_url,
                                            broadcast)

    # клавиатура (может быть None)
    keyboard = build_keyboard_for_broadcast(broadcast)

    if photo_url:
        # варианты 1 и 2:
        # 1) только картинка (caption=None)
        # 2) картинка + текст (caption есть)
        status, data = send_user_photo_via_bot(
            test_chat_id=test_chat_id,
            bot_token=bot_token,
            photo_url=photo_url,
            caption=caption,
            keyboard=keyboard,
            disable_link_preview=broadcast.disable_link_preview,
            parse_mode="HTML" if caption else None,
        )
    else:
        # вариант 3: только текст
        send_message_telegram(
            test_chat_id, text_only, bot_token,
            keyboard=keyboard,
            disable_link_preview=broadcast.disable_link_preview,
            parse_mode="HTML",  # для summernote
            )
    return f"Sent test message ID {broadcast_id}."


def send_broadcast_test_task_singleflow(broadcast_id: int):
    """Последоватльная отправка сообщения."""
    broadcast, bot, bot_token, city = get_broadcast_bot_token_city(broadcast_id)

    # выбираем тестовый админский чат по городу
    #test_chat_id = settings.ADMIN_CHATS.get(city) if city else None

    # выбираем личный чат бота с админом по боту
    test_chat_id = bot.admin_id
    if not test_chat_id or not bot_token:
        return f"No test chat or bot token for city={city}"

    cleaned_message = clean_html_for_telegram(broadcast.body)
    keyboard = build_keyboard_for_broadcast(broadcast)
    send_message_telegram(test_chat_id, cleaned_message, bot_token,
                          disable_link_preview=broadcast.disable_link_preview,
                          keyboard=keyboard,
                          parse_mode="HTML",  # для summernote
                          )

    # _send_broadcast_summary_to_admin(broadcast, bot,
    #                                  Counter({
    #                                     "ok": 1,
    #                                     "bot was blocked": 0,
    #                                     "chat not found": 0,
    #                                     "invalid chat": 0,
    #                                     "error": 0,
    #                                     "exception": 0,
    #                                  }),
    #                                  1, 1)

    return f"Sent test message ID {broadcast_id}."


@shared_task(
    queue="broadcast",
    bind=True)
def send_broadcast_task(self, broadcast_id: int):

    # 1) короткая транзакция только для "замка", чтобы не запустилось несколько воркеров
    with transaction.atomic():
        broadcast = PromoBroadcast.objects.select_for_update().get(pk=broadcast_id)

        if broadcast.status == PromoBroadcast.Status.SENDING:
            logger.warning("Broadcast %s ALREADY SENDING", broadcast.id)
            return "already sending"

        if broadcast.status == PromoBroadcast.Status.DONE:
            logger.warning("Broadcast %s ALREADY DONE", broadcast.id)
            return "already done"

        broadcast.status = PromoBroadcast.Status.SENDING
        broadcast.sent_at = timezone.now()
        broadcast.processed_count = 0
        broadcast.delivered_count = 0
        broadcast.results_json = {}
        broadcast.save(update_fields=["status","sent_at","processed_count","delivered_count","results_json"])

    # close_old_connections()
    broadcast, bot, bot_token, city = get_broadcast_bot_token_city(broadcast_id)

    # картинка (может быть None)
    photo_url = get_broadcast_photo_url(broadcast)

    # текст сообщения
    # очищаем HTML форматирование в сообщении
    cleaned_message = clean_html_for_telegram(broadcast.body)
    # обрезаем кол-во символов под стандарт и если есть фото
    caption, text_only = crop_text_if_photo(cleaned_message,
                                            photo_url,
                                            broadcast)

    # клавиатура (может быть None)
    keyboard = build_keyboard_for_broadcast(broadcast)

    # фильтруем подписчиков по городу и мессенджеру
    # отбрасываем аккаунты, где бот подходит, но у него стоит False,
    # т.е. пользователь заблочил
    # условие "этому боту можно писать"
    can_write_q = (
        Q(bot_links__bot=bot) &
        (Q(bot_links__tg_can_write=True) | Q(bot_links__tg_can_write__isnull=True))
    )

    qs = (
        MessengerAccount.objects
        .filter(msngr_type="tm", subscription=True)
        # получатели: либо в городе рассылки, либо бот может писать
        .filter(Q(city=city) | can_write_q)
        # но если по этому боту точно нельзя писать — исключаем
        .exclude(bot_links__bot=bot, bot_links__tg_can_write=False)
        .distinct()
    )

    ids = list(qs.values_list("id", flat=True))
    total = len(ids)
    broadcast.total_recipients = total
    broadcast.status = PromoBroadcast.Status.SENDING
    broadcast.sent_at = timezone.now()
    broadcast.processed_count = 0
    broadcast.delivered_count = 0
    broadcast.results_json = {}
    broadcast.save(update_fields=["total_recipients", "status", "sent_at",
                                  "processed_count", "delivered_count",
                                  "results_json"])

    stats = Counter({
        "ok": 0,
        "bot was blocked": 0,
        "chat not found": 0,
        "user is deactivated": 0,
        "invalid chat": 0,
        "rate limited": 0,
        "temporary error": 0,
        "error": 0,
        "exception": 0,
    })
    processed = 0
    delivered = 0
    # завести список строк отчёта
    detail_rows = []

    for offset in range(0, total, SAVE_EVERY):
        batch_ids = ids[offset: offset + SAVE_EVERY]

        # Важно: берём объекты отдельным запросом, курсор не живёт "вечно"
        accounts = MessengerAccount.objects.filter(id__in=batch_ids).order_by("id")

        for acc in accounts:
            final_status = None

            for attempt in range(3):  # до 3 попыток отправки
                try:

                    mab_before = MessengerAccountBot.objects.filter(
                        messenger_account=acc,
                        bot=bot,
                    ).first()

                    if photo_url:
                        status, data = send_user_photo_via_bot(
                            messenger_account=acc,
                            bot=bot,
                            bot_token=bot_token,
                            photo_url=photo_url,
                            caption=caption,
                            keyboard=keyboard,
                            disable_link_preview=broadcast.disable_link_preview,
                            parse_mode="HTML" if caption else None,
                        )
                    else:
                        if not text_only:
                            status, data = "ok", {}
                        else:
                            status, data = send_user_message_via_bot(
                                messenger_account=acc,
                                bot=bot,
                                message=text_only,
                                bot_token=bot_token,
                                keyboard=keyboard,
                                disable_link_preview=broadcast.disable_link_preview,
                                parse_mode="HTML",
                            )
                    mab_after = MessengerAccountBot.objects.filter(
                        messenger_account=acc,
                        bot=bot,
                    ).first()

                    # делаем запись о результате каждой отправки
                    detail_rows.append({
                        "messenger_account_id": acc.id,
                        "username": acc.msngr_username,
                        "msngr_id": acc.msngr_id,
                        "tm_chat_id": acc.tm_chat_id,
                        "city": acc.city,
                        "bot_id": bot.id if bot else None,
                        "bot_name": str(bot) if bot else None,
                        "status": status,
                        "error_code": data.get("error_code") if isinstance(data, dict) else None,
                        "description": data.get("description") if isinstance(data, dict) else None,
                        "response_json": json.dumps(data, ensure_ascii=False) if isinstance(data, dict) else str(data),
                        "tg_can_write_before": mab_before.tg_can_write if mab_before else None,
                        "tg_can_write_after": mab_after.tg_can_write if mab_after else None,
                        "last_error_code_after": mab_after.last_error_code if mab_after else None,
                    })
                    logger.warning("Broadcast debug: user=%s status=%s data=%s",
                                   acc.id, status, data)

                    final_status = status

                    if status == "ok":
                        delivered += 1
                        break

                    if status in ["bot was blocked", "chat not found",
                                  "invalid chat", "user is deactivated"]:
                        break

                    logger.warning(
                        "Broadcast: unexpected status '%s' for user %s via bot %s.",
                        status,
                        acc.id,
                        bot.id,
                    )
                    break

                except Exception:
                    final_status = "exception"
                    if attempt < 2:
                        time.sleep(2 ** attempt)
                        continue

                    logger.exception(
                        "Broadcast: fatal error for user %s via bot %s.",
                        acc.id,
                        bot.id,
                    )
                    break

            if final_status is None:
                final_status = "exception"

            # ✅ защита: если прилетел неожиданный статус — считаем как error
            if final_status not in stats:
                logger.warning(
                    "Broadcast: unknown status '%s' for user %s (count as error).",
                    final_status, acc.id
                )
                final_status = "error"

            stats[final_status] += 1
            processed += 1

            # лучше логгером, а не print (print в docker иногда режется)
            logger.warning("Broadcast progress: tried=%s delivered=%s",
                           processed, delivered)
            time.sleep(0.05)

        # ✅ Сохраняем прогресс ОДИН РАЗ после батча (это безопаснее)
        try:
            PromoBroadcast.objects.filter(id=broadcast.id).update(
                processed_count=processed,
                delivered_count=delivered,
                results_json=dict(stats),
            )
        except (InterfaceError, OperationalError):
            logger.warning("Broadcast %s: DB error while saving progress.",
                           broadcast.id)

    PromoBroadcast.objects.filter(id=broadcast.id).update(
        processed_count=processed,
        delivered_count=delivered,
        results_json=dict(stats),
        status=PromoBroadcast.Status.DONE,
    )
    logger.warning("Broadcast %s summary: %s", broadcast.id, dict(stats))

    broadcast.refresh_from_db(fields=[
        "status",
        "processed_count",
        "delivered_count",
        "results_json",
        "total_recipients",
        "sent_at",
    ])

    filename, content = build_broadcast_report_file(
        broadcast=broadcast,
        bot=bot,
        city=city,
        counter=stats,
        detail_rows=detail_rows,
        processed_count=processed,
        delivered_count=delivered,
        final_status=PromoBroadcast.Status.DONE,
    )

    broadcast.report_file.save(filename, content, save=False)
    broadcast.save(update_fields=["report_file"])

    _send_broadcast_summary_to_admin(broadcast, bot, dict(stats),
                                     processed, delivered)
    return delivered


def get_broadcast_bot_token_city(broadcast_id: int):
    broadcast = PromoBroadcast.objects.get(pk=broadcast_id)
    bot = broadcast.bot
    # город и токен читаем из OrdersBot (названия полей подставь свои)
    city = getattr(bot, "city", None)
    bot_token = settings.TELEGRAM_AUTH_BOTS.get(city)
    return broadcast, bot, bot_token, city


def _send_broadcast_summary_to_admin(broadcast, bot, stats,
                                     processed, delivered):
    """Отправляет итоговую статистику рассылки в админский чат бота."""
    from tm_bot.services import get_chat_id_by_bot, send_message_telegram

    admin_chat_id = get_chat_id_by_bot(bot)
    if not admin_chat_id:
        logger.warning("Broadcast %s: no admin chat, summary not sent.",
                       broadcast.id)
        return

    total = broadcast.total_recipients or processed

    def pct(n):
        return f"{n / total * 100:.1f}%" if total else "—"

    blocked = stats.get('bot was blocked', 0)
    not_started = stats.get('chat not found', 0)
    deactivated = stats.get('user is deactivated', 0)
    invalid = stats.get('invalid chat', 0)
    rate_limited = stats.get('rate limited', 0)
    temporary = stats.get('temporary error', 0)
    errors = stats.get('error', 0) + stats.get('exception', 0)

    message = (
        f"📣 <b>Рассылка завершена</b>\n"
        f"📋 {broadcast.title}\n"
        f"\n"
        f"👥 Получателей: <b>{total}</b>\n"
        f"✅ Доставлено: <b>{delivered}</b> ({pct(delivered)})\n"
        f"\n"
        f"<b>Детали:</b>\n"
        f"🚫 Заблокировал бота: <b>{blocked}</b>\n"
        f"👻 Не запускал бота: <b>{not_started}</b>\n"
        f"💀 Деактивирован: <b>{deactivated}</b>\n"
        f"❌ Нет chat_id: <b>{invalid}</b>\n"
        f"⏳ Rate limited: <b>{rate_limited}</b>\n"
        f"🌐 Временные ошибки: <b>{temporary}</b>\n"
        f"⚠️ Ошибки: <b>{errors}</b>"
    )

    try:
        send_message_telegram(
            chat_id=admin_chat_id,
            message=message,
            bot_token=settings.ADMIN_BOT_TOKEN,
            parse_mode="HTML",
            disable_link_preview=True,
        )
    except Exception:
        logger.exception("Broadcast %s: failed to send summary to admin chat.",
                         broadcast.id)

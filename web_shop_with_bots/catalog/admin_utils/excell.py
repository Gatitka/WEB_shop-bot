# catalog/reports/excell.py

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from catalog.models import Dish, DishCityPrice, DishPartnerPrice

# Размер пачки для bulk_create/bulk_update — чтобы не упереться
# в лимиты БД/памяти на больших файлах.
BATCH_SIZE = 1000

# Структура файла: строка 1 — названия городов (объединённые ячейки),
# строка 2 — подзаголовки колонок, данные начинаются с строки 3.
CITY_NAME_ROW = 1
SUBHEADER_ROW = 2
DATA_START_ROW = 3
COLS_PER_CITY_BLOCK = 4

REQUIRED_SUBHEADERS = {"базовая цена", "финальная цена", "P1", "P2"}


class ExcelImportError(Exception):
    """Критическая ошибка, прерывающая импорт целиком (например, неизвестный город
    или неполный набор колонок)."""


@dataclass
class RowError:
    row: int
    article: str
    city: str | None
    message: str


@dataclass
class ImportResult:
    updated_site: int = 0
    created_site: int = 0
    updated_partner: int = 0
    created_partner: int = 0
    skipped_no_article: int = 0
    skipped_no_dish: int = 0
    errors: list[RowError] = field(default_factory=list)

    @property
    def skipped(self) -> int:
        return self.skipped_no_article + self.skipped_no_dish

    def as_dict(self):
        return {
            "updated_site": self.updated_site,
            "created_site": self.created_site,
            "updated_partner": self.updated_partner,
            "created_partner": self.created_partner,
            "skipped_no_article": self.skipped_no_article,
            "skipped_no_dish": self.skipped_no_dish,
            "skipped": self.skipped,
            "errors": [
                {
                    "row": e.row,
                    "article": e.article,
                    "city": e.city,
                    "message": e.message,
                }
                for e in self.errors
            ],
        }


def _to_decimal(value):
    """Преобразует значение ячейки в Decimal с двумя знаками после запятой.
    Пустая ячейка -> None. Некорректное значение -> ValueError.
    """
    if value in ("", None):
        return None
    try:
        return Decimal(str(value).replace(",", ".")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        raise ValueError(f"Некорректное число: {value!r}")


def _get_article(value):
    if not value:
        return None
    return str(value).split("/")[0].strip()


def _cell(row, idx):
    """Безопасно получить значение из кортежа строки (короткие строки -> None)."""
    return row[idx] if idx < len(row) else None


def _build_city_blocks(row1, row2, allowed_cities):
    """Определяет блоки колонок по двум строкам заголовков.
    row1 / row2 — кортежи значений строк (из iter_rows(values_only=True)),
    индексация с нуля (колонка B = index 1).
    """
    city_blocks = []
    col_idx = 1  # колонка B (0 = A — артикул/название)

    while col_idx < len(row1):
        city = row1[col_idx]

        if not city:
            col_idx += 1
            continue

        city = str(city).strip()

        if city not in allowed_cities:
            raise ExcelImportError(f"Неизвестный город в файле: {city!r}")

        sub_headers = {}
        for offset in range(COLS_PER_CITY_BLOCK):
            idx = col_idx + offset
            name = _cell(row2, idx)
            if name:
                sub_headers[str(name).strip()] = idx

        missing = REQUIRED_SUBHEADERS - set(sub_headers)
        if missing:
            raise ExcelImportError(
                f"Для города {city} нет колонок: {', '.join(sorted(missing))}"
            )

        city_blocks.append({
            "city": city,
            "site_price_idx": sub_headers["базовая цена"],
            "site_final_idx": sub_headers["финальная цена"],
            "p1_idx": sub_headers["P1"],
            "p2_idx": sub_headers["P2"],
        })

        col_idx += COLS_PER_CITY_BLOCK

    return city_blocks


def import_prices_from_excel(file) -> ImportResult:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    allowed_cities = {code for code, _ in settings.CITY_CHOICES}

    # --- 1. Читаем заголовки (строки 1-2) ---
    header_rows = list(ws.iter_rows(min_row=CITY_NAME_ROW, max_row=SUBHEADER_ROW, values_only=True))

    if len(header_rows) < 2:
        wb.close()
        raise ExcelImportError("Файл не содержит обеих строк заголовков (город / колонки).")

    row1, row2 = header_rows
    city_blocks = _build_city_blocks(row1, row2, allowed_cities)

    if not city_blocks:
        wb.close()
        raise ExcelImportError("В файле не найдено ни одного известного города.")

    # --- 2. Читаем данные последовательно одним проходом (быстро в read_only режиме) ---
    result = ImportResult()
    parsed_rows = []  # (row_number, article, row_values)

    for offset, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True)):
        row_number = DATA_START_ROW + offset
        article = _get_article(_cell(row, 0))

        if not article:
            result.skipped_no_article += 1
            continue

        parsed_rows.append((row_number, article, row))

    wb.close()

    articles = {article for _, article, _ in parsed_rows}

    # --- 3. Подгружаем справочники одним запросом каждый ---
    dishes = {
        dish.article: dish
        for dish in Dish.objects.filter(article__in=articles)
    }

    # Связь идёт через related-поле dish__article, а не через dish_id напрямую,
    # т.к. dish_id — это PK модели Dish, который может не совпадать с article.
    existing_site = {
        (obj.dish.article, obj.city): obj
        for obj in DishCityPrice.objects.filter(
            dish__article__in=articles
        ).select_related("dish")
    }

    existing_partner = {
        (obj.dish.article, obj.city, obj.partner_category): obj
        for obj in DishPartnerPrice.objects.filter(
            dish__article__in=articles
        ).select_related("dish")
    }

    site_to_update = []
    site_to_create = []
    partner_to_update = []
    partner_to_create = []

    # --- 4. Обрабатываем строки в памяти, без обращений к БД внутри цикла ---
    for row_number, article, row in parsed_rows:
        dish = dishes.get(article)

        if not dish:
            result.skipped_no_dish += 1
            continue

        for block in city_blocks:
            city = block["city"]

            try:
                site_price = _to_decimal(_cell(row, block["site_price_idx"]))
                site_final = _to_decimal(_cell(row, block["site_final_idx"]))
                p1 = _to_decimal(_cell(row, block["p1_idx"]))
                p2 = _to_decimal(_cell(row, block["p2_idx"]))
            except ValueError as exc:
                result.errors.append(
                    RowError(row=row_number, article=article, city=city, message=str(exc))
                )
                continue

            # --- цена на сайте ---
            if site_price is not None and site_final is not None:
                key = (article, city)
                obj = existing_site.get(key)

                if obj:
                    obj.price = site_price
                    obj.final_price = site_final
                    site_to_update.append(obj)
                else:
                    new_obj = DishCityPrice(
                        dish=dish,
                        city=city,
                        price=site_price,
                        discount=None,
                        final_price=site_final,
                    )
                    site_to_create.append(new_obj)
                    # Регистрируем сразу, чтобы повторная строка с тем же
                    # артикулом/городом в файле обновила объект, а не создала дубль.
                    existing_site[key] = new_obj

            elif site_price is not None or site_final is not None:
                result.errors.append(
                    RowError(
                        row=row_number,
                        article=article,
                        city=city,
                        message="Заполнена только одна из цен сайта (базовая/финальная), ожидаются обе.",
                    )
                )

            # --- цены партнёров ---
            for partner_category, value in (("P1", p1), ("P2", p2)):
                if value is None:
                    continue

                key = (article, city, partner_category)
                obj = existing_partner.get(key)

                if obj:
                    obj.final_price = value
                    partner_to_update.append(obj)
                else:
                    new_obj = DishPartnerPrice(
                        dish=dish,
                        city=city,
                        partner_category=partner_category,
                        final_price=value,
                    )
                    partner_to_create.append(new_obj)
                    existing_partner[key] = new_obj

    # --- 5. Сохраняем одной транзакцией, пачками ---
    with transaction.atomic():
        if site_to_create:
            DishCityPrice.objects.bulk_create(site_to_create, batch_size=BATCH_SIZE)

        if site_to_update:
            DishCityPrice.objects.bulk_update(
                site_to_update,
                ["price", "final_price"],
                batch_size=BATCH_SIZE,
            )

        if partner_to_create:
            DishPartnerPrice.objects.bulk_create(partner_to_create, batch_size=BATCH_SIZE)

        if partner_to_update:
            DishPartnerPrice.objects.bulk_update(
                partner_to_update,
                ["final_price"],
                batch_size=BATCH_SIZE,
            )

    result.updated_site = len(site_to_update)
    result.created_site = len(site_to_create)
    result.updated_partner = len(partner_to_update)
    result.created_partner = len(partner_to_create)

    # после загрузки новых цен инвалидируем кэш
    if site_to_create or site_to_update or partner_to_create or partner_to_update:
        from api.utils.core_cache import invalidate_menu_cache
        invalidate_menu_cache()

    return result


def export_prices_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prices"

    cities = [city[0] for city in settings.CITY_CHOICES]

    ws.cell(row=1, column=1, value="Блюда")
    ws.cell(row=2, column=1, value="Артикул / Название")

    col = 2
    city_start_cols = {}

    for city in cities:
        city_start_cols[city] = col

        ws.merge_cells(
            start_row=1,
            start_column=col,
            end_row=1,
            end_column=col + 3,
        )
        ws.cell(row=1, column=col, value=city)

        ws.cell(row=2, column=col, value="базовая цена")
        ws.cell(row=2, column=col + 1, value="финальная цена")
        ws.cell(row=2, column=col + 2, value="P1")
        ws.cell(row=2, column=col + 3, value="P2")

        col += 4

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF3F8")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for row in (1, 2):
        for cell in ws[row]:
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill if row == 1 else subheader_fill

    dishes = (
        Dish.objects
        .all()
        .prefetch_related(
            "translations",
            "city_prices",
            "partner_prices",
        )
        .order_by("article")
    )

    row = 3

    for dish in dishes:
        name = dish.safe_translation_getter(
            "short_name",
            language_code="ru",
            any_language=True,
        ) or ""

        ws.cell(row=row, column=1, value=f"{dish.article} / {name}")

        city_prices = {
            price.city: price
            for price in dish.city_prices.all()
        }

        partner_prices = {
            (price.city, price.partner_category): price
            for price in dish.partner_prices.all()
        }

        for city in cities:
            start_col = city_start_cols[city]

            site = city_prices.get(city)
            p1 = partner_prices.get((city, "P1"))
            p2 = partner_prices.get((city, "P2"))

            ws.cell(row=row, column=start_col, value=site.price if site else None)
            ws.cell(row=row, column=start_col + 1, value=site.final_price if site else None)
            ws.cell(row=row, column=start_col + 2, value=p1.final_price if p1 else None)
            ws.cell(row=row, column=start_col + 3, value=p2.final_price if p2 else None)

        row += 1

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 35

    for col_idx in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = 'attachment; filename="dish_prices.xlsx"'

    wb.save(response)
    return response

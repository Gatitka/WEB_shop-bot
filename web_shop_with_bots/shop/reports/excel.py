from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook

from .periods import get_range_period, get_file_data
from .querysets import get_filtered_orders_qs, get_filtered_orderdishes_qs
from .rows import (
    build_full_order_row,
    build_full_orders_headers,
    build_order_item_row,
    build_order_items_headers,
    build_short_order_row,
    build_short_orders_headers,
)


def create_excel_response(filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def write_sheet_with_rows(workbook, title, first_row, headers, rows):
    ws = workbook.create_sheet(title=title)
    ws.append([first_row])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return ws


def export_full_orders_to_excel(modeladmin, request, queryset):
    start_date, start_pref, end_date, end_pref = get_range_period(request)
    current_date = datetime.now(timezone.utc).strftime('%d-%m-%Y')

    filename, ws_title, first_row = get_file_data(
        start_date,
        end_date,
        current_date,
        'LARGE',
    )

    response = create_excel_response(filename)
    admin = request.user

    orders_qs = get_filtered_orders_qs(
        start_date,
        start_pref,
        end_date,
        end_pref,
        admin,
    )
    orderdishes_qs = get_filtered_orderdishes_qs(
        start_date,
        start_pref,
        end_date,
        end_pref,
        admin,
    )

    wb = Workbook(write_only=True)

    write_sheet_with_rows(
        workbook=wb,
        title=ws_title[:31],
        first_row=first_row,
        headers=build_full_orders_headers(),
        rows=(
            build_full_order_row(order)
            for order in orders_qs.iterator(chunk_size=500)
        ),
    )

    write_sheet_with_rows(
        workbook=wb,
        title='Order_items',
        first_row=first_row,
        headers=build_order_items_headers(),
        rows=(
            build_order_item_row(item)
            for item in orderdishes_qs.iterator(chunk_size=1000)
        ),
    )

    wb.save(response)
    return response


export_full_orders_to_excel.short_description = (
    'Сохранить ПОЛНЫЙ отчет по продажам в Excel.'
)


def export_orders_to_excel(modeladmin, request, queryset):
    start_date, start_pref, end_date, end_pref = get_range_period(request)
    current_date = datetime.now(timezone.utc).strftime('%d-%m-%Y')

    filename, ws_title, first_row = get_file_data(
        start_date,
        end_date,
        current_date,
        'SHORT',
    )

    response = create_excel_response(filename)
    admin = request.user
    orders_qs = get_filtered_orders_qs(
        start_date,
        start_pref,
        end_date,
        end_pref,
        admin,
    )

    wb = Workbook(write_only=True)

    write_sheet_with_rows(
        workbook=wb,
        title=ws_title[:31],
        first_row=first_row,
        headers=build_short_orders_headers(),
        rows=(
            build_short_order_row(order)
            for order in orders_qs.iterator(chunk_size=500)
        ),
    )

    wb.save(response)
    return response


export_orders_to_excel.short_description = (
    'Сохранить отчет по продажам в Excel.'
)

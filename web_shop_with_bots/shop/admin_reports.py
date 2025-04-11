from django.conf import settings
from decimal import Decimal
from datetime import datetime, timedelta
from django.utils import timezone
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Q
from shop.models import Order
from django.conf import settings


def get_report_data(orders_list):
    #Разбираем заказы по типам для дальнейшего анализа
    delivery_orders = []
    takeaway_orders = []
    partners_orders = []
    for order in orders_list:
        if order.delivery.type == 'delivery':
            delivery_orders.append(order)
        elif order.delivery.type == 'takeaway':
            takeaway_orders.append(order)
            if order.source in settings.PARTNERS_LIST:
                partners_orders.append(order)

    # Calculate the total discounted amount and total receipts
    total_amount = sum(order.final_amount_with_shipping for order in orders_list)
    total_qty = orders_list.count()
    total_receipts = sum(order.invoice for order in orders_list)

    total_nocash = (sum(order.final_amount_with_shipping for order in orders_list if order.source != 'P2-2' and order.payment_type == 'cash' and order.invoice is False)
                    + sum(order.final_amount_with_shipping for order in partners_orders if order.source == 'P2-2'))    # не та дверь платит налом без чека
    total_gotovina = sum(order.final_amount_with_shipping for order in orders_list if order.payment_type == 'cash' and order.invoice is True)
    _takeaway_gotovina = sum(order.final_amount_with_shipping for order in takeaway_orders if order.payment_type == 'cash' and order.invoice is True)
    # Calculate total takeaways
    takeaway_nocash = (sum(order.final_amount_with_shipping for order in takeaway_orders if order.source != 'P2-2' and order.payment_type == 'cash' and order.invoice is False)
                       + sum(order.final_amount_with_shipping for order in partners_orders if order.source == 'P2-2'))    # не та дверь платит налом без чека
    takeaway_gotovina = total_gotovina
    takeaway_card = sum(order.final_amount_with_shipping for order in takeaway_orders if order.payment_type in ['card', 'card_on_delivery'])

    # Prepare partners data
    SOURCE_DICT = dict(settings.SOURCE_TYPES)
    partners = {}
    for order in partners_orders:
        partner_name = SOURCE_DICT[order.source]
        if partner_name in partners:
            partners[partner_name] += order.final_amount_with_shipping
        else:
            partners[partner_name] = order.final_amount_with_shipping

    total_smoke = partners.get('Smoke', Decimal('0'))
    total_curiers_show = sum(order.final_amount_with_shipping for order in delivery_orders if order.payment_type == 'cash')
    total_curiers = sum(order.final_amount_with_shipping for order in delivery_orders if order.payment_type == 'cash' and order.invoice is True)
    total_terminal = total_amount - total_nocash - total_smoke      #   - total_curiers

    curiers = get_couriers_data(delivery_orders)

    total_cash = get_cash_report_total(curiers, takeaway_nocash, _takeaway_gotovina)
    drugo_bezgotovinsko = get_bezgotovinsko_report_total(curiers, partners)

    report = {
        'total_amount': f"{total_amount:.2f} ({total_qty} зак.)",
        'takeaway_nocash': float(takeaway_nocash),
        'takeaway_gotovina': float(takeaway_gotovina),
        'takeaway_card': takeaway_card,
        'total_curiers': total_curiers_show,
        'total_terminal': total_terminal,
        #'total_receipts': total_receipts,
        'partners': partners,
        'couriers': curiers,
        'total_cash': total_cash,
        'drugo_bezgotovinsko': drugo_bezgotovinsko,
    }

    return report


def get_couriers_data(delivery_orders):
    """
    Get courier-related data aggregated by restaurant.
    couriers = {
        'courier_name': [Decimal('0') - сумма доставок для оплаты курьеру,   0
                         Bool - есть ли "уточнить",                         1
                         Decimal('0') - сумма заказов безнал,                2
                         Decimal('0') - сумма заказов нал,                  3
                         Decimal('0') - сумма заказов безнал + нал,          4
                         Decimal('0') - сумма заказов карта (безготовинско), 5
                         Decimal('0') - сумма минимальной оплаты за выход], 6
        'total_cash': Dec,
        'total_bezgotovinsko': Dec,
        }
    """

    if not delivery_orders:
        return {'Нет курьеров': [0, False, 0, 0, 0, 0, 0]}

    couriers = {}

    for order in delivery_orders:
        courier_name = order.courier if order.courier else 'Unknown'
        unclarified = False

        if order.delivery_zone.delivery_cost != float(0):
            delivery_cost = order.delivery_zone.delivery_cost
        elif order.delivery_zone.name == 'уточнить':
            delivery_cost = order.delivery_cost
            unclarified = True
        elif order.delivery_zone.name == 'по запросу':
            delivery_cost = order.delivery_cost

        if courier_name in couriers:
            couriers[courier_name][0] -= delivery_cost
        else:
            couriers[courier_name] = [Decimal('0'), False,
                                      Decimal('0'), Decimal('0'), Decimal('0'),
                                      Decimal('0'),
                                      Decimal('0')]
            if order.courier:
                couriers[courier_name][6] = order.courier.min_payout

            couriers[courier_name][0] = 0 - delivery_cost
        couriers[courier_name][1] = unclarified

        if order.payment_type == 'cash' and order.invoice is False:
            couriers[courier_name][2] += order.final_amount_with_shipping
            couriers[courier_name][4] += order.final_amount_with_shipping  # доб в тотал нал + безнал
        elif order.payment_type == 'cash' and order.invoice is True:
            couriers[courier_name][3] += order.final_amount_with_shipping
            couriers[courier_name][4] += order.final_amount_with_shipping  # доб в тотал нал + безнал
        elif order.payment_type in ['card', 'card_on_delivery']:
            couriers[courier_name][5] += order.final_amount_with_shipping

    total_cash = Decimal('0')
    total_bezgotovinsko = Decimal('0')

    # к стоимостям доставок добавляем оплату минимальную за выход
    for results in couriers.values():
        results[0] -= results[6]   # для получения полной ЗП прибавляем мин оклад

        total_cash += results[0]
        total_cash += results[4]
        total_bezgotovinsko += results[5]

    couriers.update({'total_cash': total_cash,
                     'total_bezgotovinsko': total_bezgotovinsko})

    return couriers


def get_cash_report_total(curiers, takeaway_nocash, takeaway_gotovina):
    total_cash = Decimal('0')

    # суммируем курьеров
    if 'total_cash' in curiers:
        total_cash += curiers['total_cash']
    # прибавляем безнал, gotovina
    total_cash += takeaway_nocash
    total_cash += takeaway_gotovina

    return total_cash


def get_bezgotovinsko_report_total(curiers, partners):
    drugo_bezgotovinsko = Decimal('0')

    # суммируем курьеров
    if 'total_bezgotovinsko' in curiers:
        drugo_bezgotovinsko += curiers['total_bezgotovinsko']
    # прибавляем партнеров
    for partner, total_value in partners.items():
        if partner in ['Glovo', 'Wolt']:
            drugo_bezgotovinsko += total_value

    return drugo_bezgotovinsko


#--------------------EXCELL-------------------------


def get_range_period(request):
    start_date_data = request.GET.get('execution_date__gte')
    if start_date_data is not None:
        start_date = datetime.strptime(start_date_data,
                                       '%Y-%m-%d %H:%M:%S%z')
    else:
        start_date_data = request.GET.get('execution_date__range__gte')
        if start_date_data is not None:
            start_date = datetime.strptime(start_date_data,
                                           '%d.%m.%Y')
    start_pref = 'gte'

    end_date_data = request.GET.get('execution_date__lt')
    if end_date_data is not None:
        end_date = datetime.strptime(end_date_data,
                                     '%Y-%m-%d %H:%M:%S%z')
        end_pref = 'lt'
    else:
        end_date_data = request.GET.get('execution_date__range__lte')
        if end_date_data is not None:
            end_date = (datetime.strptime(end_date_data,
                                          '%d.%m.%Y')
                        + timedelta(days=1) - timedelta(seconds=1))
            end_pref = 'lte'

    if (start_date_data is None
            and end_date_data is None):
        start_date, start_pref, end_date, end_pref = None, None, None, None

    return start_date, start_pref, end_date, end_pref


def get_file_data(start_date, end_date, current_date, type):
    if (start_date is not None
            and end_date is not None):
        start_date_str = datetime.strftime(start_date, '%d.%m.%Y')
        end_date_str = datetime.strftime(end_date, '%d.%m.%Y')
        filename = (f"{type}_orders_{start_date_str}-"
                    f"{end_date_str}_crtd_at_"
                    f"{current_date}.xlsx")
        ws_title = f"Orders_{start_date_str}-{end_date_str}"
        first_row = f"Период заказов: {start_date_str} - {end_date_str}"

    else:
        filename = (f"orders_ALL_crtd_at_"
                    f"{current_date}.xlsx")
        ws_title = "Orders_ALL"
        first_row = "Период заказов: все заказы"
    return filename, ws_title, first_row


def get_filtered_orders_qs(start_date, start_pref, end_date, end_pref, admin):
    """Если не переданы рамки временные, то берем все заказы"""
    filter_q = Q()
    if start_date is not None and end_date is not None:
        if end_pref == 'lt':
            filter_q &= Q(execution_date__gte=start_date) & Q(execution_date__lt=end_date)
        elif end_pref == 'lte':
            filter_q &= Q(execution_date__gte=start_date) & Q(execution_date__lte=end_date)

    if admin.restaurant:
        filter_q &= Q(restaurant=admin.restaurant)

    qs = Order.objects.filter(filter_q).select_related(
                'user',
                'delivery',
                'delivery_zone',
                'promocode',
                'restaurant',
                'courier'
            ).prefetch_related(
                'orderdishes__dish',
                'orderdishes__dish__translations'
            )

    return qs


def export_full_orders_to_excel(modeladmin, request, queryset):
    start_date, start_pref, end_date, end_pref = get_range_period(request)
    current_date = datetime.now(timezone.utc).strftime('%d-%m-%Y')

    filename, ws_title, first_row = get_file_data(start_date, end_date,
                                                  current_date, 'LARGE')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    admin = request.user
    queryset = get_filtered_orders_qs(start_date, start_pref,
                                      end_date, end_pref, admin)

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active

    ws.title = ws_title

    # Вставляем первую строку с периодом заказов
    ws.insert_rows(1)
    ws.cell(row=1, column=1).value = first_row

    # Заголовки столбцов
    ws.append(['Order_Num', 'ID', 'Source',
               'Date', 'Time',
               'Status', 'Is_first_order', 'Created_by',
               'Source',
               'User', 'Name',
               'Phone',
               'MSNGR_ID', 'MSNGR_USERNAME',
               'Delivery',
               'Delivery_Time',
               'Address', 'Delivery_Zone',
               'Courier',
               'Payment',
               'Invoice',
               'Discount', 'Delivery_Cost',
               'Discount amount',
               'Manual_discount', 'Amount',
               'Discounted_amount', 'Final_amount_with_shipping',
               ])

    # Добавляем данные из queryset
    for order in queryset:

        if order.is_first_order:
            is_first_order = 'yes'
        else:
            is_first_order = ''

        if order.invoice:
            invoice = 1
        else:
            invoice = ''

        ws.append(
            [
                order.order_number, order.id, str(order.source),
                order.created.astimezone(None).strftime('%Y-%m-%d'),
                order.created.astimezone(None).strftime('%H:%M:%S'),
                order.status, is_first_order, order.created_by,
                order.source,
                str(order.user), order.recipient_name,
                str(order.recipient_phone),
                order.msngr_account.msngr_id if order.msngr_account is not None else '',
                order.msngr_account.msngr_username if order.msngr_account is not None else '',
                order.delivery.type,
                (order.delivery_time.astimezone(None).strftime(
                    '%Y-%m-%d %H:%M:%S') if order.delivery_time else None),
                order.recipient_address, str(order.delivery_zone),
                str(order.courier),
                order.payment_type,
                invoice,
                str(order.discount), order.delivery_cost, order.discount_amount,
                order.manual_discount,
                order.amount,
                order.discounted_amount, order.final_amount_with_shipping,
            ]
        )

    # Сохраняем файл в HttpResponse
    wb.save(response)
    return response


export_full_orders_to_excel.short_description = (
    "Сохранить ПОЛНЫЙ отчет по продажам в Excel.")


def export_orders_to_excel(modeladmin, request, queryset):
    start_date, start_pref, end_date, end_pref = get_range_period(request)
    current_date = datetime.now(timezone.utc).strftime('%d-%m-%Y')

    filename, ws_title, first_row = get_file_data(start_date, end_date,
                                                  current_date, 'SHORT')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    admin = request.user
    queryset = get_filtered_orders_qs(start_date, start_pref,
                                      end_date, end_pref, admin)

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active

    ws.title = ws_title

    # Вставляем первую строку с периодом заказов
    ws.insert_rows(1)
    ws.cell(row=1, column=1).value = first_row

    # Заголовки столбцов
    ws.append(['Заказ',
               'Адрес',
               'Сумма',
               'N',
               'Чек',
               'Примечание',
               'Стоимость доставки',
               'Статус',
               'Курьер',
               ])

    # Добавляем данные из queryset
    for order in queryset:
        if order.delivery.type == 'delivery':
            address = order.recipient_address
        else:
            if order.source in ['1', '3', '4']:
                address = 'самовывоз'
            elif order.source == 'P1-1':
                address = "GLOVO"
            elif order.source == 'P1-2':
                address = "WOLT"
            elif order.source == 'P2-1':
                address = "SMOKE"
            elif order.source == 'P2-2':
                address = "NE TA"
            elif order.source == 'P3-1':
                address = "SEAL TEA"

        if order.invoice:
            invoice = 1
        else:
            invoice = ''

        note = order.source_id if order.source in ['3'] + settings.PARTNERS_LIST else ''

        courier = str(order.courier) if order.courier is not None else ""

        ws.append(
            [
                order.order_number,
                address,
                order.final_amount_with_shipping,
                order.payment_type,
                invoice,
                note,
                order.delivery_cost,
                order.status,
                courier
            ]
        )

    # Сохраняем файл в HttpResponse
    wb.save(response)
    return response


export_orders_to_excel.short_description = (
    "Сохранить отчет по продажам в Excel.")
